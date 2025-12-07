import pandas as pd
import os
import logging
import math
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# --- PDF IMPORTS ---
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas 
# --- END PDF IMPORTS ---

# --- Configuration ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("seating_arrangement.log", mode='w'),
        logging.StreamHandler()
    ]
)
error_logger = logging.getLogger('ErrorLogger')
error_handler = logging.FileHandler('errors.txt', mode='w')
error_handler.setLevel(logging.ERROR)
error_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
error_logger.addHandler(error_handler)

# --- Main Class ---

class SeatingArrangementPlanner:
    def __init__(self, buffer, arrangement_type, excel_file_path):
        self.buffer = buffer
        self.arrangement_type = arrangement_type.lower()
        self.excel_file = excel_file_path

        # Output directories
        self.output_dir_excel = "Exam_Seating_Arrangement"
        os.makedirs(self.output_dir_excel, exist_ok=True)
        
        self.output_dir_pdf = "Attendance_Sheets_pdf"
        os.makedirs(self.output_dir_pdf, exist_ok=True)
        
        # Photos configuration
        self.photos_dir = "photos"
        self.pic_png_path = os.path.join(self.photos_dir, 'pic.png')
        self.nopic_png_path = os.path.join(self.photos_dir, 'nopic.png')

        # Warning checks
        if not os.path.exists(self.photos_dir):
             os.makedirs(self.photos_dir, exist_ok=True)
             logging.warning(f"Created '{self.photos_dir}' folder. Please populate it.")

        if not os.path.exists(self.nopic_png_path) and os.path.exists(self.photos_dir):
            logging.warning(f"Required 'nopic.png' not found in '{self.photos_dir}'.")
        if not os.path.exists(self.pic_png_path):
            logging.warning(f"Required 'pic.png' not found in '{self.photos_dir}'.")

        self.df_schedule = None
        self.df_enrollment = None
        self.df_names = None
        self.df_rooms = None
        self.master_allocation_list = []
        self.master_seats_left_list = []

    def _clean_columns(self, df):
        cols = df.columns
        new_cols = [str(c).lower().strip().replace('.', '') for c in cols]
        rename_map = {
            'roll': 'roll number', 'rollno': 'roll number',
            'course_code': 'course code', 'course code': 'course code',
            'name': 'name',
            'room no': 'room number', 'room number': 'room number',
            'exam capacity': 'capacity', 'capacity': 'capacity',
            'block': 'building', 'building': 'building',
            'morning': 'subjects in the morning', 'subjects in the morning': 'subjects in the morning',
            'evening': 'subjects in the evening', 'subjects in the evening': 'subjects in the evening',
            'floor': 'floor'
        }
        final_cols = []
        for c in new_cols:
            mapped = False
            for key, val in rename_map.items():
                if key in c:
                    final_cols.append(val)
                    mapped = True
                    break
            if not mapped:
                final_cols.append(c)
        df.columns = final_cols
        return df

    def load_data(self):
        logging.info(f"Loading data from: {self.excel_file}")
        try:
            xls = pd.ExcelFile(self.excel_file)
            self.df_schedule = pd.read_excel(xls, sheet_name=0, header=0)
            self.df_enrollment = pd.read_excel(xls, sheet_name=1, header=0)
            self.df_names = pd.read_excel(xls, sheet_name=2, header=0)
            self.df_rooms = pd.read_excel(xls, sheet_name=3, header=0)

            self.df_schedule = self._clean_columns(self.df_schedule)
            self.df_enrollment = self._clean_columns(self.df_enrollment)
            self.df_names = self._clean_columns(self.df_names)
            self.df_rooms = self._clean_columns(self.df_rooms)
            
            self.df_schedule['date'] = pd.to_datetime(self.df_schedule['date'])
            self.df_rooms = self.df_rooms.loc[:, ~self.df_rooms.columns.str.contains('^unnamed')]
            self.df_rooms = self.df_rooms.dropna(how='all') 
            self.df_rooms['capacity'] = pd.to_numeric(self.df_rooms['capacity'], errors='coerce')
            self.df_rooms = self.df_rooms.dropna(subset=['capacity'])
            self.df_rooms['capacity'] = self.df_rooms['capacity'].astype(int)

            return True
        except Exception as e:
            logging.error(f"Data loading error: {e}")
            return False

    def get_subject_data_and_check_clashes(self, subject_codes):
        subjects_data = {}
        for code in subject_codes:
            try:
                rolls = self.df_enrollment[self.df_enrollment['course code'] == code]['roll number'].astype(str).tolist()
                if not rolls: continue
                sorted_rolls = sorted(rolls)
                subjects_data[code] = {'roll_numbers': sorted_rolls, 'count': len(sorted_rolls)}
            except KeyError:
                return None

        roll_counts = defaultdict(list)
        for code in subjects_data:
            for roll in subjects_data[code]['roll_numbers']:
                roll_counts[roll].append(code)

        has_clash = False
        for roll, courses in roll_counts.items():
            if len(courses) > 1:
                has_clash = True
        
        if has_clash: return None
        return subjects_data

    def allocate_students(self, subjects_data):
        try:
            session_rooms = self.df_rooms.copy()
            session_rooms['effective_capacity'] = session_rooms['capacity'] - self.buffer
            session_rooms.loc[session_rooms['effective_capacity'] < 0, 'effective_capacity'] = 0
            session_rooms['remaining_capacity'] = session_rooms['effective_capacity']
            session_rooms['allocated_subjects'] = [{} for _ in range(len(session_rooms))]
            
            sorted_subject_codes = sorted(subjects_data.keys(), key=lambda s: subjects_data[s]['count'], reverse=True)
            session_rooms = session_rooms.sort_values(by=['building', 'effective_capacity'], ascending=[True, False]).reset_index(drop=True)
            
            session_allocations = []
            total_students_to_allocate = sum(data['count'] for data in subjects_data.values())
            total_room_capacity = session_rooms['effective_capacity'].sum()

            if total_students_to_allocate > total_room_capacity:
                return None, None

            for subject_code in sorted_subject_codes:
                students_to_allocate = list(subjects_data[subject_code]['roll_numbers'])
                allocated_in_building = None
                
                while students_to_allocate:
                    available_rooms = session_rooms[session_rooms['remaining_capacity'] > 0]
                    if allocated_in_building:
                        preferred_rooms = available_rooms[available_rooms['building'] == allocated_in_building]
                        target_rooms = preferred_rooms if not preferred_rooms.empty else available_rooms
                    else:
                        target_rooms = available_rooms
                    
                    if target_rooms.empty: break
                    
                    room_index = target_rooms.index[0] 
                    room_details = session_rooms.loc[room_index]
                    
                    if not allocated_in_building:
                        allocated_in_building = room_details['building']

                    if self.arrangement_type == 'sparse':
                        max_per_subject = math.floor(room_details['effective_capacity'] / 2)
                        already_in_room = room_details['allocated_subjects'].get(subject_code, 0)
                        capacity_for_slot = max(0, min(room_details['remaining_capacity'], max_per_subject - already_in_room))
                    else:
                        capacity_for_slot = room_details['remaining_capacity']

                    if capacity_for_slot <= 0:
                        session_rooms.loc[room_index, 'remaining_capacity'] = -999 
                        continue

                    num_to_place = min(len(students_to_allocate), int(capacity_for_slot))
                    placed_students = students_to_allocate[:num_to_place]
                    students_to_allocate = students_to_allocate[num_to_place:]
                    
                    session_rooms.loc[room_index, 'remaining_capacity'] -= num_to_place
                    current_allocs = session_rooms.at[room_index, 'allocated_subjects']
                    current_allocs[subject_code] = current_allocs.get(subject_code, 0) + num_to_place
                    
                    session_allocations.append({
                        'course_code': subject_code,
                        'Room': str(room_details['room number']),
                        'Building': room_details['building'],
                        'Allocated_students_count': num_to_place,
                        'Roll_list': placed_students
                    })
            return session_allocations, session_rooms
        except Exception as e:
            logging.error(f"Allocation Error: {e}")
            return None, None

    def generate_outputs(self, date_info, session_name, allocations, final_room_state):
        try:
            date_str_iso = date_info.strftime('%Y-%m-%d')
            date_str_pdf = date_info.strftime('%d-%m-%Y') 
            day_str = date_info.strftime('%A')
            
            session_folder_excel = os.path.join(self.output_dir_excel, f"{date_str_iso}_{day_str}", session_name)
            os.makedirs(session_folder_excel, exist_ok=True)
            
            session_folder_pdf = os.path.join(self.output_dir_pdf, f"{date_str_iso}_{day_str}", session_name)
            os.makedirs(session_folder_pdf, exist_ok=True)
            
            for alloc in allocations:
                self.master_allocation_list.append({
                    'Date': date_str_iso, 'Day': day_str, 'Session': session_name, 
                    'Building': alloc['Building'], **alloc
                })
                
                pdf_filename = f"{date_str_iso.replace('-', '_')}_{session_name.replace(' ', '')}_{alloc['Room']}_{alloc['course_code']}.pdf"
                pdf_filepath = os.path.join(session_folder_pdf, pdf_filename)
                
                self.create_pdf_attendance_sheet(
                    path=pdf_filepath, date_str=date_str_pdf, session_name=session_name,
                    subject_code=alloc['course_code'], room_number=alloc['Room'],
                    roll_numbers=alloc['Roll_list'], total_students=alloc['Allocated_students_count'],
                    day_name=day_str
                )

                filename = f"{date_str_iso}_{alloc['course_code']}_{alloc['Room']}.xlsx"
                filepath = os.path.join(session_folder_excel, filename)
                self.create_attendance_sheet(
                    path=filepath, date_str=date_str_pdf, session_name=session_name,
                    subject_code=alloc['course_code'], room_number=alloc['Room'],
                    roll_numbers=alloc['Roll_list']
                )

            for _, room in final_room_state.iterrows():
                total_allotted = room['effective_capacity'] - room['remaining_capacity'] if room['remaining_capacity'] > -990 else room['effective_capacity']
                self.master_seats_left_list.append({
                    'Date': date_str_iso, 'Session': session_name,
                    'Room No.': str(room['room number']), 'Block': room['building'],
                    'Exam Capacity': room['capacity'], 'Alloted': total_allotted,
                    'Vacant (B-C)': room['capacity'] - total_allotted
                })
        except Exception as e:
            logging.error(f"Output Generation Error: {e}")

    def create_pdf_attendance_sheet(self, path, date_str, session_name, subject_code, room_number, roll_numbers, total_students, day_name):
        try:
            self.df_names['roll number'] = self.df_names['roll number'].astype(str)
            rolls_df = pd.DataFrame(roll_numbers, columns=['roll number'])
            
            student_details = pd.merge(rolls_df, self.df_names, on='roll number', how='left')
            student_details['name'].fillna('(name not found)', inplace=True) 
            student_details = student_details[['roll number', 'name']].reset_index(drop=True)
            
            doc = SimpleDocTemplate(path, pagesize=A4, 
                                    rightMargin=0.3*inch, leftMargin=0.3*inch, 
                                    topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = []
            
            styles = getSampleStyleSheet()
            style_left = ParagraphStyle('Header_Left', parent=styles['Normal'], alignment=TA_LEFT, fontSize=10, leading=12)
            style_center_bold = ParagraphStyle('Center_Bold', parent=styles['Normal'], alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=10, spaceBefore=10, spaceAfter=5)

            def get_image_flowable(roll):
                student_img_path = os.path.join(self.photos_dir, f"{roll}.jpg") 
                img_dim = 0.6 * inch 
                try:
                    if os.path.exists(student_img_path):
                        if os.path.exists(self.pic_png_path):
                            return Image(self.pic_png_path, width=img_dim, height=img_dim)
                        else:
                            return Paragraph("Img Found", styles['Normal'])
                    elif os.path.exists(self.nopic_png_path):
                        return Image(self.nopic_png_path, width=img_dim, height=img_dim)
                    else:
                        return Paragraph("No Image", styles['Normal'])
                except Exception:
                    return Paragraph("Err", styles['Normal'])

            def first_page_header(canvas, doc):
                canvas.saveState()
                
                title = "IITP Attendance System"
                canvas.setFont('Helvetica-Bold', 18)
                title_w = canvas.stringWidth(title, 'Helvetica-Bold', 18)
                canvas.drawString((A4[0] - title_w) / 2, A4[1] - 0.75 * inch, title)

                date_text = f"<b>Date:</b> {date_str} ({day_name}) | <b>Shift:</b> {session_name} | <b>Room No:</b> {room_number} | <b>Student count:</b> {total_students}"
                p_date = Paragraph(date_text, style_left)
                
                subj_text = f"<b>Subject:</b> {subject_code}"
                pres_text = "<b>Stud Present:</b>"
                abs_text = "<b>| Stud Absent:</b>"
                
                p_subj = Paragraph(subj_text, style_left)
                p_pres = Paragraph(pres_text, style_left)
                p_abs = Paragraph(abs_text, style_left)
                
                header_data = [
                    [p_date, '', ''], 
                    [p_subj, p_pres, p_abs]
                ]
                
                header_table = Table(header_data, colWidths=[3.5*inch, 2.0*inch, 2.1*inch])
                header_table.setStyle(TableStyle([
                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                    ('SPAN', (0,0), (-1,0)),
                    ('LEFTPADDING', (0,0), (-1,-1), 5),
                    ('RIGHTPADDING', (0,0), (-1,-1), 0),
                    ('TOPPADDING', (0,0), (-1,-1), 2),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ]))
                
                w, h = header_table.wrap(doc.width, doc.topMargin)
                header_table.drawOn(canvas, doc.leftMargin, A4[1] - 1.6 * inch) 
                canvas.restoreState()

            story.append(Spacer(0, 1.6 * inch))
            
            table_data = []
            students_per_row = 3
            num_students = len(student_details)
            
            for i in range(0, num_students, students_per_row):
                row = []
                for j in range(students_per_row):
                    student_index = i + j
                    if student_index < num_students:
                        student = student_details.iloc[student_index]
                        img_flowable = get_image_flowable(student['roll number'])
                        p_name = Paragraph(f"<b>{student['name']}</b>", styles['Normal'])
                        p_roll = Paragraph(f"<b>Roll:</b> {student['roll number']}", styles['Normal'])
                        p_sign = Paragraph("<b>Sign:</b> ____________", styles['Normal'])
                        
                        block_content = [[img_flowable, [p_name, p_roll, p_sign]]]
                        block_table = Table(block_content, colWidths=[0.7*inch, 1.6*inch])
                        block_table.setStyle(TableStyle([
                            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                            ('LEFTPADDING', (0, 0), (-1, -1), 2),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                            ('TOPPADDING', (0, 0), (-1, -1), 2),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                        ]))
                        row.append(block_table)
                    else:
                        row.append("") 
                table_data.append(row)

            col_width = 2.5 * inch
            main_table = Table(table_data, colWidths=[col_width, col_width, col_width])
            main_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(main_table)
            
            story.append(Spacer(0, 0.4 * inch))
            story.append(Paragraph("<b>Invigilator Name & Signature</b>", style_center_bold))
            
            invig_data = [["SI No.", "Name", "Signature"]] + [["", "", ""] for _ in range(5)]
            
            invig_table = Table(invig_data, colWidths=[0.8*inch, 3.5*inch, 3*inch], rowHeights=0.3*inch)
            invig_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'), 
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]))
            story.append(invig_table)

            doc.build(story, onFirstPage=first_page_header, onLaterPages=lambda canvas, doc: None) 
            logging.info(f"Successfully created PDF: {path}")

        except Exception as e:
            logging.error(f"FAILURE: Could not create PDF at {path}. Error: {e}", exc_info=True)

    def create_attendance_sheet(self, path, date_str, session_name, subject_code, room_number, roll_numbers):
        try:
            self.df_names['roll number'] = self.df_names['roll number'].astype(str)
            rolls_df = pd.DataFrame(roll_numbers, columns=['roll number'])
            student_details = pd.merge(rolls_df, self.df_names, on='roll number', how='left')
            student_details['name'].fillna('Unknown Name', inplace=True) 
            student_details = student_details[['roll number', 'name']]
            student_details.columns = ['Roll Number', 'Student Name']
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            header_font = Font(bold=True, size=14)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.merge_cells('A1:B1')
            ws['A1'] = f"Course: {subject_code} | Room: {room_number} | Date: {date_str} | Session: {session_name}"
            ws['A1'].font = header_font
            ws['A1'].alignment = center_align
            
            rows = dataframe_to_rows(student_details, index=False, header=True)
            for r_idx, row in enumerate(rows, 3):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(path)
        except Exception:
            pass

    def finalize_reports(self):
        try:
            if self.master_allocation_list:
                df = pd.DataFrame(self.master_allocation_list)
                df['Roll_list (semicolon separated_,'] = df['Roll_list'].apply(lambda r: ';'.join(map(str, r)))
                df.to_excel(os.path.join(self.output_dir_excel, "op_overall_seating_arrangement.xlsx"), index=False)
            
            if self.master_seats_left_list:
                df = pd.DataFrame(self.master_seats_left_list).drop_duplicates()
                df.to_excel(os.path.join(self.output_dir_excel, "op_seats_left.xlsx"), index=False)
        except Exception as e:
            logging.error(f"Error finalizing: {e}")

    def run(self):
        if not self.load_data(): return
        for _, row in self.df_schedule.iterrows():
            date_info = row['date']
            for session in ['morning', 'evening']:
                col = f'subjects in the {session}'
                if col not in row or pd.isna(row[col]) or 'no exam' in str(row[col]).lower(): continue
                
                subjects = [s.strip() for s in str(row[col]).split(';') if s.strip()]
                data = self.get_subject_data_and_check_clashes(subjects)
                if data:
                    allocs, state = self.allocate_students(data)
                    if allocs: self.generate_outputs(date_info, session.capitalize(), allocs, state)
        self.finalize_reports()

def main():
    print("--- Exam Seating Arrangement Generator (Docker CLI) ---")
    
    excel_file = 'input_data_tt.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"Error: '{excel_file}' not found in the current directory.")
        return

    try:
        buffer_input = input("Enter seat buffer (e.g., 5): ")
        buffer = int(buffer_input)
        
        arrangement_type = ""
        while arrangement_type not in ['sparse', 'dense']:
            arrangement_type = input("Enter arrangement type ('sparse' or 'dense'): ").lower()

        planner = SeatingArrangementPlanner(
            buffer=buffer, 
            arrangement_type=arrangement_type,
            excel_file_path=excel_file
        )
        print("Processing... (Check log files if it takes time)")
        planner.run()
        print("\nProcess Completed. Check 'Exam_Seating_Arrangement' and 'Attendance_Sheets_pdf' folders.")

    except ValueError:
        print("Invalid input. Buffer must be an integer.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()